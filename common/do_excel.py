"""
@Time    : 2019/4/29 0029 16:26
@Author  : Cooper
@File    : do_excel.py
@Function: 完成excel的读取
"""
import openpyxl
from common.do_log import MyLog


class Case:
    def __init__(self):
        self.case_id = None
        self.case_name = None
        self.function_name = None
        self.url = None
        self.data = None
        self.expected = None
        self.actul = None
        self.result = None
        self.sql = None


class Do_excel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        try:
            self.wb = openpyxl.load_workbook(self.file_name)
        except Exception as e:
            MyLog(__name__).my_log().error('文件不存在或文件路径错误错误')
        self.sh = self.wb[self.sheet_name]

    def get_data(self):
        cases = []
        for row in range(2, self.sh.max_row + 1):
            case = Case()
            case.case_id = self.sh.cell(row=row, column=1).value
            case.case_name = self.sh.cell(row=row, column=2).value
            case.function_name = self.sh.cell(row=row, column=3).value
            case.url = self.sh.cell(row=row, column=4).value
            case.data = self.sh.cell(row=row, column=5).value
            case.expected = self.sh.cell(row=row, column=6).value
            case.sql = self.sh.cell(row=row, column=9).value
            cases.append(case)
        self.wb.close()
        return cases

    def write_data(self, row, actual, result):
        sheet = self.wb[self.sheet_name]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.wb.save(self.file_name)
        self.wb.close()


if __name__ == '__main__':
    from common import contants

    a = Do_excel(contants.case_file, 'sendMCode')
    b = a.get_data()
    for i in b:
        print(i.__dict__)
